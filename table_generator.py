#!/usr/bin/env python3
# -*- coding=utf-8 -*-
'''
This script is used to generate an xlsx table from the report html index
'''
import csv
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
from report_process import get_studentId_percentage

classdict = dict()

def load_student_dict(csvfile):
    '''
    load a mapping from id to name
    '''
    filed = open(csvfile)
    csv_reader = csv.reader(filed)
    for row in csv_reader:
        classdict[row[0]] = row[1]

def get_name_from_id(stud_id):
    if not stud_id or stud_id == '':
        return 'Not found'
    try:
        return classdict[stud_id]
    except KeyError:
        return 'Not found'

def insert_name_row(row):
    '''
    insert the name into the row according to the mapping
    '''
    stud_id_list = [row[1], row[3]]
    stud_name = []
    for stud_id in stud_id_list:
        stud_name.append(get_name_from_id(stud_id))
    row.insert(4, stud_name[1])
    row.insert(2, stud_name[0])


def read_from_html(htmlfile):
    '''
    read data from html and return the tabledata
    '''
    tabledata = list()
    file = open(htmlfile)
    htmltext = file.read()
    file.close()
    soup = BeautifulSoup(htmltext, 'lxml')
    allrows = soup.find_all("tr")[1:]
    for row in allrows:
        rowdata = list()
        tds = row.find_all("td")
        for td in tds[0:2]:
            rowdata += list(get_studentId_percentage(td.a.text))
        rowdata.append(tds[2].text.strip())
        # print(tds[2].text)
        tabledata.append(rowdata)
    return tabledata


def init_table(tablefile):
    '''
    A: labnum
    B: stuId1
    C: percentage1
    D: stuId2
    E: percentage2
    F: matched lines
    '''
    wb = Workbook()
    ws1 = wb.active
    tableheader = ['LabNum', 'StuId1', 'StuName1', 'Percentage1/%',
                   'StuId2', 'StuName2', 'Percentage2/%', 'Matched lines']
    for idx, header in enumerate(tableheader):
        ws1.cell(row=1, column=idx + 1, value=header)
    wb.save(tablefile)


def append_table(tablefile, tabledata, labnum):
    '''
    append data
    generate xlsx table form the data given
    '''
    wb = load_workbook(tablefile)
    ws1 = wb.active
    row_offset = ws1.max_row
    for rowidx, row in enumerate(tabledata):
        row.insert(0, labnum)
        insert_name_row(row)
        for columnidx, info in enumerate(row):
            # print(info)
            ws1.cell(row=row_offset + rowidx + 1,
                     column=columnidx + 1, value=info)
    wb.save(tablefile)


def main():
    load_student_dict("./data/classDict.csv")
    init_table("./judge/report/result.xlsx")
    for idx in range(1, 10):
        tabledata = read_from_html("./judge/report/lab%d/index.html" % idx)
        append_table("./judge/report/result.xlsx", tabledata, idx)


if __name__ == '__main__':
    main()
